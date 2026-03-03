"""
Утилиты: форматированный вывод, стилизация ячеек, работа с openpyxl.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
import os


def print_header(text):
    """Печатает заголовок с рамкой."""
    print("\n" + "=" * 60)
    print(f"   {text}")
    print("=" * 60)


def print_success(text):
    """Печатает сообщение об успехе."""
    print(f"✅ {text}")


def print_error(text):
    """Печатает сообщение об ошибке."""
    print(f"❌ {text}")


def print_info(text):
    """Печатает информационное сообщение."""
    print(f"ℹ️ {text}")


def apply_header_style(worksheet, cell_range):
    """Применяет стиль заголовка к диапазону ячеек."""
    for row in worksheet[cell_range]:
        for cell in row:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")


def apply_borders(worksheet, cell_range):
    """Применяет границы к диапазону ячеек."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = thin_border


def auto_adjust_column_width(worksheet):
    """Автоматически подгоняет ширину колонок под содержимое."""
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width


def safe_remove_file(filename):
    """Безопасно удаляет файл, если он существует."""
    if os.path.exists(filename):
        os.remove(filename)
        print_info(f"Старый файл '{filename}' удалён.")
