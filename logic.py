"""
Модуль содержит логику каждой демонстрации.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from core import generate_sample_employees, generate_sample_sales
from utils import (
    print_header, print_success, print_info, print_error,
    apply_header_style, apply_borders, auto_adjust_column_width, safe_remove_file
)
import datetime

FILENAME = "demo_workbook.xlsx"


def demo_create_simple_table():
    """1. Создание простой таблицы с данными."""
    print_header("📄 1. СОЗДАНИЕ ПРОСТОЙ ТАБЛИЦЫ (ЗАПИСЬ ДАННЫХ)")

    # Генерируем данные
    employees = generate_sample_employees(5)

    # Создаём книгу и активный лист
    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    # Заголовки
    headers = ["ID", "ФИО", "Отдел", "Зарплата", "Дата приёма"]
    ws.append(headers)

    # Данные
    for emp in employees:
        ws.append([emp.id, emp.name, emp.department, emp.salary, emp.hire_date])

    # Стилизация
    apply_header_style(ws, "A1:E1")
    apply_borders(ws, f"A1:E{len(employees)+1}")
    auto_adjust_column_width(ws)

    # Сохраняем
    safe_remove_file(FILENAME)
    wb.save(FILENAME)

    print_success(f"Файл '{FILENAME}' создан с листом 'Сотрудники'.")
    print_info("Содержимое файла:")
    for emp in employees:
        print(f"   {emp.id} | {emp.name} | {emp.department} | {emp.salary} | {emp.hire_date}")


def demo_read_data():
    """2. Чтение данных из существующего Excel файла."""
    print_header("📖 2. ЧТЕНИЕ ДАННЫХ ИЗ EXCEL")

    try:
        wb = load_workbook(FILENAME)
        ws = wb.active
        print_info(f"Активный лист: {ws.title}")

        print_info("Первые 5 строк:")
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            print(f"   Строка {i}: {row}")
            if i >= 6:
                break

        # Покажем количество строк с данными
        data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(row))
        print_success(f"Всего строк с данными: {data_rows}")
    except FileNotFoundError:
        print_error(f"Файл '{FILENAME}' не найден. Сначала выполните демонстрацию 1.")
    except Exception as e:
        print_error(f"Ошибка при чтении: {e}")


def demo_append_data():
    """3. Добавление новых данных в существующий файл."""
    print_header("➕ 3. ДОБАВЛЕНИЕ НОВЫХ ДАННЫХ")

    try:
        wb = load_workbook(FILENAME)
        ws = wb.active

        # Генерируем ещё одного сотрудника
        new_emp = generate_sample_employees(1)[0]
        ws.append([new_emp.id + 100, new_emp.name, new_emp.department, new_emp.salary, new_emp.hire_date])

        # Обновляем границы для новой строки
        last_row = ws.max_row
        apply_borders(ws, f"A{last_row}:E{last_row}")
        auto_adjust_column_width(ws)

        wb.save(FILENAME)
        print_success(f"Добавлен новый сотрудник: {new_emp.name}")
    except FileNotFoundError:
        print_error("Сначала создайте файл (демо 1).")
    except Exception as e:
        print_error(f"Ошибка: {e}")


def demo_formulas():
    """4. Работа с формулами (подсчёт суммы, среднего)."""
    print_header("🧮 4. РАБОТА С ФОРМУЛАМИ")

    try:
        wb = load_workbook(FILENAME)
        ws = wb.active

        # Добавляем строку с формулами
        last_row = ws.max_row
        formula_row = last_row + 2

        ws[f"A{formula_row}"] = "Итого:"
        ws[f"B{formula_row}"] = "=SUM(D2:D{})".format(last_row)
        ws[f"C{formula_row}"] = "Среднее:"
        ws[f"D{formula_row}"] = "=AVERAGE(D2:D{})".format(last_row)

        # Стилизация для строки формул
        ws[f"A{formula_row}"].font = ws[f"B{formula_row}"].font = ws[f"C{formula_row}"].font = ws[f"D{formula_row}"].font = Font(bold=True)
        apply_borders(ws, f"A{formula_row}:D{formula_row}")

        wb.save(FILENAME)
        print_success("Добавлены формулы SUM и AVERAGE для колонки зарплат.")
    except FileNotFoundError:
        print_error("Сначала создайте файл.")
    except Exception as e:
        print_error(f"Ошибка: {e}")


def demo_formatting():
    """5. Форматирование ячеек (цвет, шрифт, выравнивание)."""
    print_header("🎨 5. ФОРМАТИРОВАНИЕ ЯЧЕЕК")

    try:
        wb = load_workbook(FILENAME)
        ws = wb.active

        from openpyxl.styles import PatternFill, Font, Alignment

        # Применим разные стили к существующим данным
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column == 4:  # колонка зарплаты
                    if cell.value and isinstance(cell.value, (int, float)):
                        if cell.value > 100000:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif cell.value < 70000:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif cell.column == 2:  # ФИО
                    cell.font = Font(italic=True)

        # Выравнивание по центру для всех ячеек
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        wb.save(FILENAME)
        print_success("Применено условное форматирование (зарплата >100k зелёным, <70k красным).")
        print_success("ФИО выделены курсивом, все ячейки отцентрированы.")
    except FileNotFoundError:
        print_error("Сначала создайте файл.")


def demo_chart():
    """6. Создание диаграммы."""
    print_header("📊 6. СОЗДАНИЕ ДИАГРАММЫ")

    try:
        wb = load_workbook(FILENAME)
        ws = wb.active

        # Определим данные для диаграммы (колонка зарплат)
        last_row = ws.max_row
        data = Reference(ws, min_col=4, min_row=2, max_row=last_row)
        categories = Reference(ws, min_col=2, min_row=2, max_row=last_row)

        chart = BarChart()
        chart.title = "Зарплаты сотрудников"
        chart.x_axis.title = "Сотрудники"
        chart.y_axis.title = "Зарплата"

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)

        # Разместим диаграмму справа от данных (например, начиная с колонки G)
        ws.add_chart(chart, "G2")

        wb.save(FILENAME)
        print_success("Диаграмма добавлена на лист.")
    except FileNotFoundError:
        print_error("Сначала создайте файл.")
    except Exception as e:
        print_error(f"Ошибка при создании диаграммы: {e}")


def demo_multiple_sheets():
    """7. Работа с несколькими листами (создание второго листа с продажами)."""
    print_header("📑 7. РАБОТА С НЕСКОЛЬКИМИ ЛИСТАМИ")

    try:
        wb = load_workbook(FILENAME)

        # Создадим лист с продажами
        if "Продажи" in wb.sheetnames:
            ws_sales = wb["Продажи"]
            print_info("Лист 'Продажи' уже существует, данные будут добавлены.")
        else:
            ws_sales = wb.create_sheet("Продажи")
            # Заголовки
            headers = ["ID", "Товар", "Кол-во", "Цена", "Сумма", "Дата"]
            ws_sales.append(headers)
            apply_header_style(ws_sales, "A1:F1")

        # Генерируем продажи
        sales = generate_sample_sales(5)
        for sale in sales:
            total = sale.quantity * sale.price
            ws_sales.append([sale.id, sale.product, sale.quantity, sale.price, total, sale.date])

        # Применяем границы и автоширину
        last_row = ws_sales.max_row
        apply_borders(ws_sales, f"A1:F{last_row}")
        auto_adjust_column_width(ws_sales)

        wb.save(FILENAME)
        print_success("Лист 'Продажи' создан/обновлён с данными о продажах.")
    except FileNotFoundError:
        print_error("Сначала создайте файл.")
    except Exception as e:
        print_error(f"Ошибка: {e}")
